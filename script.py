import pandas as pd
import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime
import warnings
from openpyxl.utils.exceptions import InvalidFileException
import gc
import sys
import re
import xlrd

def date_to_calendar_week(date_obj):
    """Converts a date object to a calendar week string (YYCWXX)."""
    if pd.isna(date_obj):
        return None
    year, week, _ = date_obj.isocalendar()
    return f"{year % 100:02d}CW{week:02d}"

def get_current_calendar_week():
    """Gets the current calendar week in YYCWXX format."""
    now = datetime.now()
    return date_to_calendar_week(now)

def mm_yyyy_to_yycwxx(mm_yyyy):
    """Converts MM/YYYY format to YYCWXX format."""
    try:
        month, year = map(int, mm_yyyy.split('/'))
        year = year % 100
        # Calculate the calendar week
        date_obj = datetime.strptime(f"{year:02d}-{month}-1", "%y-%m-%d")
        _, week, _ = date_obj.isocalendar()
        return f"{year:02d}CW{week:02d}"
    except (ValueError, TypeError):
        return None

def read_excel_file(input_file, sheet_name=None, header=0):
    """
    Reads an Excel file (.xls or .xlsx) and returns a pandas DataFrame.

    Args:
        input_file (str): The path to the Excel file.
        sheet_name (str, optional): The name of the sheet to read. Defaults to None (first sheet).
        header (int, optional): The row number to use as the header. Defaults to 0.

    Returns:
        pandas.DataFrame: The DataFrame containing the data from the Excel file, or None if an error occurs.
    """
    try:
        with warnings.catch_warnings(record=True) as w:
            warnings.simplefilter("ignore", category=UserWarning)
            try:
                if input_file.endswith(".xls"):
                    with pd.ExcelFile(input_file, engine="xlrd") as xlsx:
                        df = pd.read_excel(xlsx, sheet_name=sheet_name, header=header)
                else:
                    with pd.ExcelFile(input_file, engine="openpyxl") as xlsx:
                        df = pd.read_excel(xlsx, sheet_name=sheet_name, header=header)
            except InvalidFileException:
                print(f"Warning: Could not open file '{input_file}'. It might be corrupted or not a valid Excel file.")
                return None
            except Exception as e:
                print(f"Warning: An error occurred while opening '{input_file}': {e}")
                return None

            for warning in w:
                if "Workbook contains no default style" not in str(warning.message):
                    print(f"Warning in file {input_file}: {warning.message}")
        return df
    except Exception as e:
        print(f"Warning: An unexpected error occurred while reading '{input_file}': {e}")
        return None

def extract_data_from_excel_ma(input_file):
    """Extracts data for Project MA."""
    df = read_excel_file(input_file)
    if df is None:
        return None

    try:
        # Find relevant columns (handle KeyError)
        try:
            customer_item_col = df.columns.get_loc("Customer Item")
            quantity_col = df.columns.get_loc("Quantity")
            planned_receipt_date_col = df.columns.get_loc("Planned Receipt Date")
        except KeyError:
            print(f"Warning: Required columns not found in '{input_file}'.")
            return None

        # Data cleaning and conversion
        df["Planned Receipt Date"] = pd.to_datetime(df["Planned Receipt Date"])
        df["Calendar Week"] = df["Planned Receipt Date"].apply(date_to_calendar_week)
        df = df.dropna(subset=["Customer Item", "Quantity", "Calendar Week"])

        # Group by Customer Item and Calendar Week, sum quantities
        grouped = df.groupby(["Customer Item", "Calendar Week"])["Quantity"].sum().reset_index()

        # Prepare data for output
        extracted_data = []
        for index, row in grouped.iterrows():
            item_data = {
                "customer_item": row["Customer Item"],
                "calendar_week": row["Calendar Week"],
                "quantity": int(row["Quantity"]),
            }
            extracted_data.append(item_data)
        
        # Explicitly delete the DataFrame and run garbage collection
        del df
        gc.collect()

        return extracted_data

    except Exception as e:
        print(f"Warning: An unexpected error occurred while processing '{input_file}': {e}")
        return None

def extract_data_from_excel_mb(input_file):
    """Extracts data for Project MB."""
    df = read_excel_file(input_file, sheet_name="Zeitraum bis Bedarfsende", header=None)
    if df is None:
        return None

    try:
        # Extract Customer Item from cell A2
        customer_item_raw = df.iloc[1, 0]
        match = re.search(r"Sachnummer:\s+(\S+)", customer_item_raw)
        if match:
            customer_item = match.group(1)
        else:
            print(f"Warning: Could not extract Customer Item from cell A2 in '{input_file}'.")
            return None

        # Extract Calendar Weeks from row 6
        calendar_weeks_raw = df.iloc[5, 1:].tolist()
        calendar_weeks = [mm_yyyy_to_yycwxx(cw) for cw in calendar_weeks_raw if pd.notna(cw)]
        
        # Extract Quantities from row 8
        quantities = df.iloc[7, 1:].tolist()

        # Prepare data for output
        extracted_data = []
        for cw, qty in zip(calendar_weeks, quantities):
            if cw is not None and pd.notna(qty):
                item_data = {
                    "customer_item": customer_item,
                    "calendar_week": cw,
                    "quantity": int(qty) if pd.notna(qty) else 0,
                }
                extracted_data.append(item_data)

        # Explicitly delete the DataFrame and run garbage collection
        del df
        gc.collect()

        return extracted_data

    except Exception as e:
        print(f"Warning: An unexpected error occurred while processing '{input_file}': {e}")
        return None

def create_output_excel(data_list, output_file):
    """
    Creates a formatted Excel file with a pivot table-like structure, showing all calendar weeks,
    freezes the first column and first row, and highlights the current calendar week cell in yellow.
    """
    if not data_list:
        print("No data to create output Excel file.")
        return

    # Get all unique calendar weeks and sort them (oldest first)
    all_calendar_weeks = sorted(list(set([item["calendar_week"] for item in data_list])))

    # Get the current calendar week
    current_cw = get_current_calendar_week()

    # Create a new workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Extracted Data"

    # Define styles
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow fill

    # Create header row
    header_row = ["Customer Item"] + all_calendar_weeks
    ws.append(header_row)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    # Highlight the current calendar week cell in yellow
    try:
        current_cw_col_index = all_calendar_weeks.index(current_cw) + 2  # +2 to account for "Customer Item" column and 0-based index
        ws.cell(row=1, column=current_cw_col_index).fill = yellow_fill
    except ValueError:
        pass  # Current calendar week not found in the data

    # Create data rows
    customer_items = sorted(list(set([item["customer_item"] for item in data_list])))
    for customer_item in customer_items:
        row_data = [customer_item]
        for cw in all_calendar_weeks:
            quantity = 0
            for item in data_list:
                if item["customer_item"] == customer_item and item["calendar_week"] == cw:
                    quantity = item["quantity"]
                    break
            row_data.append(quantity)
        ws.append(row_data)

    # Freeze the first row and first column
    ws.freeze_panes = "B2"  # Freeze panes at cell B2

    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    # Save the workbook
    wb.save(output_file)

def process_all_excel_files(project_code):
    """
    Processes all Excel files in the current directory and generates a consolidated output file.
    """
    # Use the current directory as both input and output location
    current_dir = os.path.dirname(os.path.abspath(__file__))

    # Generate a timestamp for the output filename (YYYYMMDD_HHMM)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M")
    output_file = os.path.join(current_dir, f"extracted_data_{timestamp}.xlsx")

    all_data = []
    processed_files = []

    # Get all Excel files in the current directory, excluding output files
    excel_files = [f for f in os.listdir(current_dir)
                   if (f.endswith('.xlsx') or f.endswith('.xls'))
                   and not f.startswith('extracted_data_')]

    if not excel_files:
        print("No Excel files found in the current directory.")
        return

    print(f"Found {len(excel_files)} Excel files to process.")

    # Process each file
    files_with_data = 0
    for file in excel_files:
        file_path = os.path.join(current_dir, file)
        print(f"Processing file: {file}")
        if project_code == "MA":
            file_data = extract_data_from_excel_ma(file_path)
        elif project_code == "MB":
            file_data = extract_data_from_excel_mb(file_path)
        else:
            print(f"Warning: Invalid project code '{project_code}'.")
            return

        if file_data is not None:
            files_with_data += 1
            all_data.extend(file_data)
            processed_files.append(file)
        else:
            print(f"Warning: File '{file}' was skipped due to an error.")

    # Create the output file
    if all_data:
        create_output_excel(all_data, output_file)
        print(f"Output saved to: {output_file}")
    else:
        print("No data was extracted from the Excel files.")
    
    if processed_files:
        print("\nFiles processed successfully:")
        for file in processed_files:
            print(f"- {file}")
    else:
        print("\nNo files were processed successfully.")

if __name__ == "__main__":
    print("Excel Data Extraction Tool")
    print("=" * 30)
    print("This script will process all Excel files in its directory.")
    print("The output will be saved in the same directory.")
    print("=" * 30)

    if len(sys.argv) < 2:
        print("Error: Please provide a project code (MA or MB) as the first argument.")
        sys.exit(1)

    project_code = sys.argv[1].upper()  # Convert to uppercase for case-insensitivity
    process_all_excel_files(project_code)

    # Keep console window open until user presses Enter
    input("\nPress Enter to exit...")
