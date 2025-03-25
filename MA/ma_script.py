import pandas as pd
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime
import warnings
from openpyxl.utils.exceptions import InvalidFileException
import gc

def date_to_calendar_week(date_obj):
    """Converts a date object to a calendar week string (YYCWXX)."""
    if pd.isna(date_obj):
        return None
    year, week, _ = date_obj.isocalendar()
    return f"{year % 100:02d}CW{week:02d}"

def get_current_calendar_week():
    """Gets the current calendar week in YYCWXX format."""
    now = datetime.now()
    return date_to_calendar_week(now)

def extract_data_from_excel(input_file):
    """
    Extracts data, combines rows with same customer item and calendar week, sums quantities.
    Handles potential issues with open files and forces file closure.
    """
    try:
        with warnings.catch_warnings(record=True) as w:
            warnings.simplefilter("ignore", category=UserWarning)
            try:
                # Use a context manager to ensure the file is closed
                with pd.ExcelFile(input_file, engine="openpyxl") as xlsx:
                    df = pd.read_excel(xlsx, header=0)
            except InvalidFileException:
                print(f"Warning: Could not open file '{input_file}'. It might be corrupted or not a valid Excel file.")
                return None
            except Exception as e:
                print(f"Warning: An error occurred while opening '{input_file}': {e}")
                return None
            
            for warning in w:
                if "Workbook contains no default style" not in str(warning.message):
                    print(f"Warning in file {input_file}: {warning.message}")

        # Find relevant columns (handle KeyError)
        try:
            customer_item_col = df.columns.get_loc("Customer Item")
            quantity_col = df.columns.get_loc("Quantity")
            planned_receipt_date_col = df.columns.get_loc("Planned Receipt Date")
        except KeyError:
            print(f"Warning: Required columns not found in '{input_file}'.")
            return None

        # Data cleaning and conversion
        df["Planned Receipt Date"] = pd.to_datetime(df["Planned Receipt Date"])
        df["Calendar Week"] = df["Planned Receipt Date"].apply(date_to_calendar_week)
        df = df.dropna(subset=["Customer Item", "Quantity", "Calendar Week"])

        # Group by Customer Item and Calendar Week, sum quantities
        grouped = df.groupby(["Customer Item", "Calendar Week"])["Quantity"].sum().reset_index()

        # Prepare data for output
        extracted_data = []
        for index, row in grouped.iterrows():
            item_data = {
                "customer_item": row["Customer Item"],
                "calendar_week": row["Calendar Week"],
                "quantity": int(row["Quantity"]),
            }
            extracted_data.append(item_data)
        
        # Explicitly delete the DataFrame and run garbage collection
        del df
        gc.collect()

        return extracted_data

    except Exception as e:
        print(f"Warning: An unexpected error occurred while processing '{input_file}': {e}")
        return None

def generate_calendar_weeks(start_week, end_week):
    """Generates a list of calendar weeks between start_week and end_week (inclusive)."""
    start_year = int(start_week[:2])
    start_week_num = int(start_week[4:])
    end_year = int(end_week[:2])
    end_week_num = int(end_week[4:])

    calendar_weeks = []
    current_year = start_year
    current_week_num = start_week_num

    while True:
        calendar_weeks.append(f"{current_year:02d}CW{current_week_num:02d}")
        if current_year == end_year and current_week_num == end_week_num:
            break
        current_week_num += 1
        if current_week_num > 53:
            current_week_num = 1
            current_year += 1
            if current_year > 99:
                current_year = 0
    return calendar_weeks

def post_process_calendar_weeks(data_list):
    """Inserts missing calendar weeks into the data, filling with 0 quantities."""
    if not data_list:
        return []

    # Create a DataFrame for easier manipulation
    df = pd.DataFrame(data_list)

    # Get unique customer items and calendar weeks
    customer_items = df["customer_item"].unique()
    calendar_weeks = df["calendar_week"].unique()

    # Sort calendar weeks
    calendar_weeks.sort()

    # Determine the full range of calendar weeks
    start_week = calendar_weeks[0]
    end_week = calendar_weeks[-1]
    all_calendar_weeks = generate_calendar_weeks(start_week, end_week)

    # Create a new DataFrame with all calendar weeks
    all_data = []
    for item in customer_items:
        for week in all_calendar_weeks:
            quantity = df[(df["customer_item"] == item) & (df["calendar_week"] == week)]["quantity"].iloc[0] if (df[(df["customer_item"] == item) & (df["calendar_week"] == week)]["quantity"].any()) else 0
            all_data.append({"customer_item": item, "calendar_week": week, "quantity": quantity})

    return all_data

def create_output_excel(data_list, output_file):
    """
    Creates a formatted Excel file with a pivot table-like structure, showing all calendar weeks,
    freezes the first column and first row, and highlights the current calendar week cell in yellow.
    """
    if not data_list:
        print("No data to create output Excel file.")
        return

    # Get all unique customer items and calendar weeks
    customer_items = sorted(list(set([item["customer_item"] for item in data_list])))
    all_calendar_weeks = sorted(list(set([item["calendar_week"] for item in data_list])))

    # Get the current calendar week
    current_cw = get_current_calendar_week()

    # Create a new workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Extracted Data"

    # Define styles
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow fill

    # Create header row
    header_row = ["Customer Item"] + all_calendar_weeks
    ws.append(header_row)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    # Highlight the current calendar week cell in yellow
    try:
        current_cw_col_index = all_calendar_weeks.index(current_cw) + 2  # +2 to account for "Customer Item" column and 0-based index
        ws.cell(row=1, column=current_cw_col_index).fill = yellow_fill
    except ValueError:
        pass  # Current calendar week not found in the data

    # Create data rows
    for customer_item in customer_items:
        row_data = [customer_item]
        for cw in all_calendar_weeks:
            # Find the quantity for this customer item and calendar week, or default to 0
            quantity = next((item["quantity"] for item in data_list if item["customer_item"] == customer_item and item["calendar_week"] == cw), 0)
            row_data.append(quantity)
        ws.append(row_data)

    # Freeze the first row and first column
    ws.freeze_panes = "B2"  # Freeze panes at cell B2

    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    # Save the workbook
    wb.save(output_file)

def process_all_excel_files():
    """
    Processes all Excel files in the current directory and generates a consolidated output file.
    """
    # Use the current directory as both input and output location
    current_dir = os.path.dirname(os.path.abspath(__file__))

    # Generate a timestamp for the output filename (YYYYMMDD_HHMM)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M")
    output_file = os.path.join(current_dir, f"extracted_data_{timestamp}.xlsx")

    all_data = []
    processed_files = []

    # Get all Excel files in the current directory, excluding output files
    excel_files = [f for f in os.listdir(current_dir)
                   if (f.endswith('.xlsx') or f.endswith('.xls'))
                   and not f.startswith('extracted_data_')]

    if not excel_files:
        print("No Excel files found in the current directory.")
        return

    print(f"Found {len(excel_files)} Excel files to process.")

    # Process each file
    files_with_data = 0
    for file in excel_files:
        file_path = os.path.join(current_dir, file)
        print(f"Processing file: {file}")
        file_data = extract_data_from_excel(file_path)

        if file_data is not None:
            files_with_data += 1
            all_data.extend(file_data)
            processed_files.append(file)
        else:
            print(f"Warning: File '{file}' was skipped due to an error.")

    # --- INSERTION POINT ---
    # Post-processing step: Insert missing calendar weeks
    if all_data:  # Only post-process if there's data
        all_data = post_process_calendar_weeks(all_data)
    # --- END INSERTION POINT ---

    # Create the output file
    if all_data:
        create_output_excel(all_data, output_file)
        print(f"Output saved to: {output_file}")
    else:
        print("No data was extracted from the Excel files.")
    
    if processed_files:
        print("\nFiles processed successfully:")
        for file in processed_files:
            print(f"- {file}")
    else:
        print("\nNo files were processed successfully.")

if __name__ == "__main__":
    print("Excel Data Extraction Tool")
    print("=" * 30)
    print("This script will process all Excel files in its directory.")
    print("The output will be saved in the same directory.")
    print("=" * 30)

    # Process all files in the current directory
    process_all_excel_files()

    # Keep console window open until user presses Enter
    input("\nPress Enter to exit...")
