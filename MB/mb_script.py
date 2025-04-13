import pandas as pd
import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
import re
from datetime import datetime, date
import zipfile

def get_current_calendar_week():
    """Calculates the current calendar week in the format 'WW/YYYY'."""
    now = date.today()
    year, week, _ = now.isocalendar()
    return f"{week:02d}/{year}"

def process_mb_files(input_dir, output_file, mercedes_file):
    """Processes MB files, extracts data, and updates the Mercedes file."""
    all_data = []
    all_data_bedarfs = []
    all_data_ruckstand = []
    print(f"Processing files in directory: {input_dir}")  # Add this line
    for filename in os.listdir(input_dir):
        print(f"Checking file: {filename}")  # Add this line
        if filename.endswith(('.xls', '.xlsx')) and not filename.startswith(("mb_extracted_data_", "~$")) and filename != os.path.basename(mercedes_file):
            print(f"Processing file: {filename}")  # Add this line
            filepath = os.path.join(input_dir, filename)
            try:
                df = pd.read_excel(filepath, sheet_name="Zeitraum bis Bedarfsende", header=None)
                customer_item_raw = df.iloc[1, 0]
                match = re.search(r"Sachnummer:\s+(\S+)", customer_item_raw)
                customer_item = match.group(1) if match else None
                if customer_item is None:
                    print(f"Warning: Could not extract Customer Item from {filename}")
                    continue
                print(f"Sachnummer found: {customer_item}")

                # Extract calendar weeks from row 6 (index 5)
                calendar_weeks = df.iloc[5, 1:].tolist()

                # Extract data for Bedarf
                quantities_bedarf = df.iloc[7, 1:].tolist()
                temp_data_bedarf = []
                for cw, qty in zip(calendar_weeks, quantities_bedarf):
                    if pd.notna(cw) and pd.notna(qty):
                        temp_data_bedarf.append({"customer_item": customer_item, "calendar_week": cw, "quantity": int(qty)})
                temp_data_bedarf.sort(key=lambda x: (int(x['calendar_week'].split('/')[1]), int(x['calendar_week'].split('/')[0])))
                all_data_bedarfs.extend(temp_data_bedarf)

                # Find rows with "ABS" followed by a number
                abs_rows = []
                for index, row in df.iterrows():
                    first_cell_value = row.iloc[0]
                    if isinstance(first_cell_value, str) and re.match(r"^\s*ABS\s+\d+\w*", first_cell_value):
                        abs_rows.append(index)
                if abs_rows:
                    print(f"ABS rows found in {filename}")
                else:
                    print(f"No ABS rows found in {filename}")

                # Extract data for each ABS row
                for row_index in abs_rows:
                    abs_value_raw = df.iloc[row_index, 0].strip()
                    if isinstance(abs_value_raw, str) and abs_value_raw.startswith("ABS "):
                        abs_value = abs_value_raw.split(" ", 1)[1]
                    else:
                        abs_value = abs_value_raw
                    quantities = df.iloc[row_index, 1:].tolist()

                    # Get current week index
                    current_week = get_current_calendar_week()
                    try:
                        current_week_index = calendar_weeks.index(current_week)
                    except ValueError:
                        print(f"Warning: Current week {current_week} not found in {filename}")
                        continue

                    # Extract the 5 data points
                    extracted_quantities = quantities[current_week_index:current_week_index + 5]

                    # Ensure we have 5 data points, fill with None if not enough
                    while len(extracted_quantities) < 5:
                        extracted_quantities.append(None)

                    all_data.append({
                        "customer_item": customer_item,
                        "abs_value": abs_value,
                        "quantities": extracted_quantities,
                        "calendar_weeks": [calendar_weeks[i] if i < len(calendar_weeks) else None for i in range(current_week_index, current_week_index + 5)]
                    })
                # Extract data for Rückstand
                df_bkm = pd.read_excel(filepath, sheet_name="BKM Lieferbeziehung", header=None)
                for row_index in abs_rows:
                    abs_value_raw = df.iloc[row_index, 0].strip()
                    if isinstance(abs_value_raw, str) and abs_value_raw.startswith("ABS "):
                        abs_value = abs_value_raw.split(" ", 1)[1]
                    else:
                        abs_value = abs_value_raw
                    ruckstand = df_bkm.iloc[row_index, 21]
                    all_data_ruckstand.append({
                        "customer_item": customer_item,
                        "abs_value": abs_value,
                        "ruckstand":ruckstand
                    })

            except Exception as e:
                print(f"Error processing {filename}: {e}")

    if all_data:
        update_mercedes_file(all_data, all_data_ruckstand, mercedes_file)
    if all_data_bedarfs:
        create_output_excel(all_data_bedarfs, output_file)
        print(f"Output saved to: {output_file}")
    else:
        print("No data found.")

def update_mercedes_file(data_list, data_list_ruckstand, mercedes_file):
    if not os.path.exists(mercedes_file):
        print(f"Error: Mercedes file '{mercedes_file}' not found.")
        return
    print(f"Mercedes file exists: {mercedes_file}")
    try:
        wb = load_workbook(mercedes_file)
        print(f"Mercedes file loaded successfully: {mercedes_file}")
        ws = wb["EDI"]  # Access the "EDI" sheet
    except FileNotFoundError:
        print(f"Error: Mercedes file '{mercedes_file}' not found.")
        return
    except KeyError:
        print(f"Error: Sheet 'EDI' not found in '{mercedes_file}'.")
        return
    except PermissionError:
        print(f"Error: Permission denied to access '{mercedes_file}'. Is the file open in another program?")
        return
    except zipfile.BadZipFile:
        print(f"Error: '{mercedes_file}' is corrupted or not a valid Excel file.")
        return
    except Exception as e:
        print(f"Error loading Mercedes file: {e}")
        return

    header_font = Font(bold=True)
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    # Create a dictionary to store existing data for faster lookup
    existing_data = {}
    for row in ws.iter_rows(min_row=2):
        sachnummer = row[3].value
        abs_value_raw = row[6].value
        if sachnummer and abs_value_raw:
            if isinstance(abs_value_raw, str) and abs_value_raw.startswith("ABS "):
                abs_value = abs_value_raw.split(" ", 1)[1]
                row[6].value = abs_value
            else:
                abs_value = abs_value_raw
            existing_data[(sachnummer, str(abs_value))] = row
    
    # Create a dictionary for faster lookup of Rückstand
    ruckstand_data = {}
    for item in data_list_ruckstand:
        ruckstand_data[(item["customer_item"], item["abs_value"])] = item["ruckstand"]

    for data in data_list:
        sachnummer = data["customer_item"]
        abs_value = data["abs_value"]
        quantities = data["quantities"]
        calendar_weeks = data["calendar_weeks"]

        # Find the row or create a new one
        match_key = (sachnummer, str(abs_value))
        if match_key in existing_data:
            row = existing_data[match_key]
        else:
            # Add a new row
            new_row = [None] * ws.max_column
            ws.append(new_row)
            rows_list = list(ws.rows)
            row = rows_list[-1]
            row[3+10].value = sachnummer  # Sachnummer
            row[6+9].value = abs_value  # ABS
            # Highlight the new row in yellow
            for cell in row:
                cell.fill = yellow_fill
        # Fill Rückstand
        if match_key in ruckstand_data:
            row[9+8].value = ruckstand_data[match_key]
            
        # # Clear existing data in the five columns before filling
        # columns_to_clear = [11+8, 13+8, 15+8, 17+8, 19+8]  # S U W Y AA (Corrected indices)
        # for col_index in columns_to_clear:
        #     row[col_index-1].value = None
            
        # Fill in the quantities and calendar weeks
        columns = [11+8, 13+8, 15+8, 17+8, 19+8]  # S U W Y AA (Corrected indices)
        for i, (qty, cw) in enumerate(zip(quantities, calendar_weeks)):
            if qty is not None:
                row[columns[i]-1].value = qty
            if cw is not None:
                ws.cell(row=1, column=columns[i]).value = cw
    try:
        wb.save(mercedes_file)
    except PermissionError:
        print(f"Error: Permission denied to save '{mercedes_file}'. Is the file open in another program?")
        return
    except Exception as e:
        print(f"Error saving Mercedes file: {e}")
        return


def create_output_excel(data_list, output_file):
    if not data_list:
        return

    all_calendar_weeks = sorted(list(set([item["calendar_week"] for item in data_list])), key=lambda x: (int(x.split('/')[1]), int(x.split('/')[0])))

    wb = Workbook()
    ws = wb.active
    ws.title = "Extracted Data"
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow fill

    header_row = ["Customer Item"] + all_calendar_weeks
    ws.append(header_row)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    # Define the desired order
    desired_order = [
        "A2238305705",
        "A2238305706",
        "A2068305905",
        "A2548302703",
        "A2148308201",
        "A2978306501",
        "A2979970200",
        "A2979970600",
        "A0005003700",
        "A0005003101",
        "A0005004901",
        "A0005005301",
        "A0005002901",
    ]

    # Extract customer items and sort them based on the desired order or alphabetically
    customer_items = sorted(list(set([item["customer_item"] for item in data_list])), key=lambda x: (desired_order.index(x) if x in desired_order else len(desired_order)))

    for customer_item in customer_items:
        row_data = [customer_item]
        for cw in all_calendar_weeks:
            quantity = next((item["quantity"] for item in data_list if item["customer_item"] == customer_item and item["calendar_week"] == cw), 0)
            row_data.append(quantity)
        ws.append(row_data)
    ws.freeze_panes = "B2"

    # Highlight current week
    now = date.today()
    year, week, _ = now.isocalendar()
    current_week = f"{week:02d}/{year}"

    try:
        col_index = all_calendar_weeks.index(current_week) + 2
        ws.cell(row=1, column=col_index).fill = yellow_fill
    except ValueError:
        print(f"Warning: Current week {current_week} not found in data.")

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width
    wb.save(output_file)

if __name__ == "__main__":
    input_directory = os.path.dirname(os.path.abspath(__file__))
    timestamp = datetime.now().strftime("%y%m%d_%H%M")
    output_excel_file = f"mb_extracted_data_{timestamp}.xlsx"
    mercedes_excel_file = "Mercedes_Shipping_Plan_EDI.xlsx"
    process_mb_files(input_directory, output_excel_file, mercedes_excel_file)
